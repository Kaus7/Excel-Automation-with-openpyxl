import openpyxl
from openpyxl.styles import Font

master_data = openpyxl.load_workbook("Master.xlsx")
daily_data = openpyxl.load_workbook("daily.xlsx")

master_sheet = master_data["Sheet1"]
daily_sheet = daily_data["Sheet1"]

# Row count for Daily sheet
is_data = True
daily_row_count = 1

while is_data:
    daily_row_count += 1
    data = daily_sheet.cell(row=daily_row_count, column=1).value
    if data is None:
        is_data = False

# Row count for Master sheet
is_data = True
master_row_count = 1

while is_data:
    master_row_count += 1
    data = master_sheet.cell(row=master_row_count, column=1).value
    if data is None:
        is_data = False

# get data from daily sheet
# Extract data and store it into list of dictionaries

todays_data = []

for i in range(1, daily_row_count):
    row_data = {}
    row_data['id'] = daily_sheet.cell(row=i, column=1).value
    row_data['todays_purchase'] = daily_sheet.cell(row=i, column=2).value
    row_data['todays_rewards'] = daily_sheet.cell(row=i, column=3).value
    todays_data.append(row_data)

# Write daily sheet data into master sheet
# Find row using ID
# Go to total purchase and add today's purchase
# Go to total rewards and add today's rewards

for i in range(2,master_row_count):
    id = master_sheet.cell(row=i, column=1).value
    for row in todays_data:
        if row['id'] == id :
            todays_purchase = int(row['todays_purchase'])
            todays_reward = int(row['todays_rewards'])
            # Get data from master sheet
            total_purchase = master_sheet.cell(row=i, column=6).value
            total_reward = master_sheet.cell(row=i, column=7).value

            # Add todays values to total values
            new_purchase = total_purchase+todays_purchase
            new_reward = total_reward+todays_reward

            master_sheet.cell(row=i, column=6).value = new_purchase
            master_sheet.cell(row=i, column=7).value = new_reward

master_data.save("Master.xlsx")

daily_report = openpyxl.Workbook()
ws = daily_report.active

#Get headers
is_data = True
column_count = 1
header_values = []

while is_data:
    column_count += 1
    data = master_sheet.cell(row=1, column=column_count).value
    if data != None:
        header_values.append(data)
    else:
        is_data = False

header_style = Font("Times New Roman", size= 12, bold=True)

for i, col_name in enumerate(header_values):
    col_index = i+1
    ws.cell(row=1, column=col_index).value = col_name
    ws.cell(row=1, column=col_index).font = header_style

IDs = []
for data in todays_data:
    IDs.append(data['id'])

IDs.pop(0)


final_data = []
for i in range(2,master_row_count):
    id = master_sheet.cell(row=i , column=1).value
    if id in IDs:
        lst=[]
        for j in range(2,8):
            lst.append(master_sheet.cell(row=i, column=j).value)
        final_data.append(lst)

for data in final_data:
    ws.append(data)
daily_report.save("daily_report_send.xlsx")





